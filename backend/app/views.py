from rest_framework import generics
from .models import Sensor, Ambiente, Historico
from .serializers import SensorSerializer, AmbienteSerializer, HistoricoSerializer, RegisterSerializer
from django.shortcuts import render, redirect
from rest_framework.permissions import IsAuthenticatedOrReadOnly, IsAuthenticated
from .permissions import Gestor
from rest_framework_simplejwt.views import TokenObtainPairView
from django_filters.rest_framework import DjangoFilterBackend
from .filters import SensorFilter, HistoricoFilter, AmbienteFilter
from rest_framework.views import APIView
from django.http import HttpResponse
from openpyxl import Workbook




# -------------------------- SENSOR ---------------
class SensorListCreateView(generics.ListCreateAPIView):
    queryset = Sensor.objects.all()
    serializer_class = SensorSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_class = SensorFilter
    def get_permissions(self):
        if self.request.method == 'POST':
            return [Gestor()]
        return [IsAuthenticatedOrReadOnly()]

class SensorDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Sensor.objects.all()
    serializer_class = SensorSerializer
    def get_permissions(self):
        if self.request.method in ['PUT', 'PATCH', 'DELETE']:
            return [Gestor()]
        return [IsAuthenticatedOrReadOnly()]


# -------------------- AMBIENTE ------------------
class AmbienteListCreateView(generics.ListCreateAPIView):
    queryset = Ambiente.objects.all()
    serializer_class = AmbienteSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_class = AmbienteFilter
    def get_permissions(self):
        if self.request.method == 'POST':
            return [Gestor()]
        return [IsAuthenticatedOrReadOnly()]

class AmbienteDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Ambiente.objects.all()
    serializer_class = AmbienteSerializer
    def get_permissions(self):
        if self.request.method in ['PUT', 'PATCH', 'DELETE']:
            return [Gestor()]
        return [IsAuthenticatedOrReadOnly()]


# ------------------------ HISTÓRICO -------------------
class HistoricoListCreateView(generics.ListCreateAPIView):
    queryset = Historico.objects.all()
    serializer_class = HistoricoSerializer
    filter_backends = [DjangoFilterBackend]
    filterset_class = HistoricoFilter
    def get_permissions(self):
        if self.request.method == 'POST':
            return [Gestor()]
        return [IsAuthenticatedOrReadOnly()]

class HistoricoDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Historico.objects.all()
    serializer_class = HistoricoSerializer
    def get_permissions(self):
        if self.request.method in ['PUT', 'PATCH', 'DELETE']:
            return [Gestor()]
        return [IsAuthenticatedOrReadOnly()]
    
# --------------------- USUARIO -------------------------
class RegisterView(generics.CreateAPIView):
    serializer_class = RegisterSerializer
    permission_classes = []  


# --------------------- EXPORTAR DADOS -------------------
class ExportarAmbiente(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        wb = Workbook()
        ws = wb.active
        ws.title = "Ambientes"
        queryset = Ambiente.objects.all()
        for ambiente in queryset:
            ws.append([
                ambiente.id,
                ambiente.sig,
                ambiente.descricao,
                ambiente.ni,
                ambiente.responsavel
            ])
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="ambientes.xlsx"'
        wb.save(response)
        return response

class ExportarSensor(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        wb = Workbook()
        ws = wb.active
        ws.title = "Sensores"
        queryset = Sensor.objects.all()
        for sensor in queryset:
            ws.append([
                sensor.id,
                sensor.sensor,
                sensor.mac_address,
                sensor.unidade_med,
                sensor.latitude,
                sensor.longitude,
                'Ativo' if sensor.status else 'Inativo'
            ])
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="sensores_mesclados.xlsx"'
        wb.save(response)
        return response

class ExportarHistorico(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request, *args, **kwargs):
        wb = Workbook()
        ws = wb.active
        ws.title = "Histórico"
        queryset = Historico.objects.all()
        for historico in queryset:
            ws.append([
                historico.id,
                historico.sensor.sensor,  
                historico.ambiente.sig,   
                historico.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
                historico.valor
            ])
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="historico.xlsx"'
        wb.save(response)
        return response

